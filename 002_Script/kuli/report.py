# -*- coding: utf-8 -*-


__authors__ = 'Duy Nguyen'
__contact__ = 'duykn@gcs-vn.com'


import logging
from pathlib import Path

from openpyxl import load_workbook

import config as CONF
import utils
from table import AmsCsv, Testcase
from utils import format_string as fstr


class Report(object):

    def __init__(self, path):
        self.data = self.check_coverage(path)
        title = '{func} @ {src_rel}'.format(**self.data)
        utils.set_console_title(title)

        result = Result(path, **self.data)
        self.mode = result.get_mode()
        self.files = result.collect_files()
        self.wb = None

        self.collect_files()
        self.check_tctbl()
        self.check_csv()

        self.is_ready()

    def is_ready(self):
        '''Check result files is ready for delivering'''
        ready = True
        for tbl in [self.tctbl, self.csvtbl]:
            if tbl.errors != {}:
                ready = False
                tbl.print_error()

        if ready is False:
            print('\nPlease resolve all issue to continue ...')
            utils.close()

        # Remove source if not use stub
        if self.data['stub'] == []:
            del self.files['stub']

    def check_coverage(self, path):
        logging.info("Checking test coverage '%s'", fstr(Path(path), 0))

        def get_value(line, char=':'):
            logging.debug("Parsing line %s", [line])

            index = line.index(char)
            return line[(index+1):].strip()

        def find_package(src_full):
            '''Find package base on source path'''
            path = src_full.split('\\root\\')[0]
            dirname = Path(path).name
            config = utils.load_config()

            package = None
            for pkg in config.get('packages', {}).keys():
                if dirname.startswith(pkg):
                    package = pkg

            return package

        lines = utils.read_file(path)
        keys = ['func_full', 'src_full', 'c0', 'c1', 'mcdc', 'date',
                'func', 'src_dir', 'src_name']
        data = {keys[i]: get_value(lines[i])
                for i in range(len(keys)-3)}

        # Check C0/C1/MCDC
        [int(data[k][:-1]) for k in ['c0', 'c1', 'mcdc']]

        src_rel = data['src_full'].split('\\target\\')[-1]
        data.update({
            'func': data['func_full'].split('/')[-1],
            'src_dir': Path(src_rel).parent,
            'src_name': Path(src_rel).name,
            'src_rel': src_rel,
            'summary': lines[:6],
            'package': find_package(data['src_full']),
        })

        # Print test coverage summary
        space = ' '*3
        logging.info("Test coverage summary\n\n%s%s",
                     space, space.join(lines[:6]))

        return data

    def check_tctbl(self):
        path = self.files['tbl']
        logging.info("Checking testcase table '%s'", fstr(path, 0))

        self.tctbl = Testcase(path, self.data['func'])
        self.data['confirm'] = self.tctbl.confirm

    def check_csv(self):
        path = self.files['csv']
        logging.info("Checking csv '%s'", fstr(path, 0))

        self.csvtbl = AmsCsv(path, self.data)

        lst_stub, lst_non_stub = self.csvtbl.get_stub_info()
        self.data['stub'] = lst_stub
        self.data['non_stub'] = lst_non_stub

    def collect_files(self):
        '''Collect result files'''
        logging.info("Collecting test result files ")

        for key, path in self.files.items():
            if key == 'stub':
                continue

            with open(path, errors='ignore') as _:
                logging.info("%s", fstr(path))

    def deliver_result(self, target, template=CONF.FP_TEMPLATE):
        logging.info("Delivering test result files")
        dir_spec = Path(target).joinpath(CONF.DIR_SPEC)
        dir_result = Path(target).joinpath(CONF.DIR_RESULT, self.data['func'])

        # Copy result files
        for key, path in self.files.items():
            logging.info("%s", fstr(path))

            dst = dir_result.joinpath(path.name)
            utils.copy_file(path, dst)

            self.files[key] = dst

        # Copy unit test specification
        fname = '{0}.xlsx'.format(self.data['func'])
        dst = dir_spec.joinpath(fname)
        logging.info("%s", fstr(dst))
        utils.copy_file(template, dst)
        self.files['xlsx'] = dst

    def xlsx_update(self):
        '''Update unit test specification'''
        path = self.files['xlsx']
        logging.info("Updating unit test specification '%s'", fstr(path, 0))

        self.wb = load_workbook(path)

        # Update worksheet 1
        self.xlsx_update_ws1()

        # Update worksheet 6
        self.xlsx_update_ws6()

    def xlsx_save(self):
        path = self.files['xlsx']
        self.wb.active = 1
        self.wb.save(path)

    def xlsx_update_ws1(self):
        logging.info(fstr("Updating worksheet 1"))

        self.wb.active = 1
        ws = self.wb.active

        ws['F8'] = str(self.data['src_dir']).replace('\\', '/')
        ws['F9'] = self.data['src_name']
        ws['F10'] = self.data['func']
        ws['F11'] = '{0}.csv'.format(self.data['func'])

        f14 = 'テスト結果: {confirm}\n' \
            'Ｃ０網羅率 : {c0}\n' \
            'Ｃ１網羅率 : {c1}\n' \
            'ＭＣ／ＤＣ網羅率 : {mcdc}\n'\
            '問題点 : {issue}' \
            .format(**self.data)
        ws['F14'] = f14

        # Update 1.2
        for label, text in CONF.XLS_LABEL.items():
            row, col = utils.find_cell(ws, text)
            lb_data = self.tctbl.label_data.get(label, {})

            # r, c = utils.find_cell_next(ws, row, col, count_next=1)
            if lb_data == {}:
                ws.cell(row, col+4, '不要')
            else:
                ws.cell(row, col+4, '必要')
                ws.cell(row, col+7, '-')
                # i = 0
                # for cmt, lst in lb_data.items():
                #     if i > 0:
                #         row = row + 1
                #         ws.insert_rows(row)
                #     ws.cell(row, col + 5, cmt)
                #     ws.cell(row, col + 6, utils.collapse_list(lst))
                #     i += 1

        # Save excel
        self.xlsx_save()

    def xlsx_update_ws6(self):
        logging.info(fstr("Updating worksheet 6"))

        self.wb.active = 6
        ws = self.wb.active

        lines = [l.replace('\n', '')
                 for l in utils.read_file(self.files['txt'])]

        row = 5
        for i in range(len(lines)):
            cell = 'A{0}'.format(row + i)
            ws[cell] = lines[i]

        # Save excel
        self.xlsx_save()

    def update_issue(self, info):
        '''Update issue information'''
        if (
            self.data['c0'] != '100%' or
            self.data['c1'] != '100%' or
            self.data['mcdc'] != '100%'
        ):
            self.data['confirm'] = 'NG'

        if self.data['confirm'] == 'NG':
            number = input('\nFound <number> bugs: ')
            number = int(number)
            lst = ['{0}.{1}'.format(info['func_no'], i+1)
                   for i in range(number)]
            issue = ', '.join(lst)
            issue = '{0}No{1}'.format(info['issue_prefix'], issue)

        else:
            issue = 'なし'

        self.data['issue'] = issue


class Result(object):

    config = {
        'tbl': (2, '', '_Table.html'),
        'csv': (3, 'TestCsv', '.csv'),
        'ini': (3, 'TestCsv', '.ini'),
        'xeat': (3, 'TestCsv', '.xeat'),
        'xtct': (3, 'TestCsv', '.xtct'),
        'ie': (3, 'TestCsv', '_IE.html'),
        'io': (3, 'TestCsv', '_IO.html'),
        'oe': (3, 'TestCsv', '_OE.html'),
        'tc': (3, 'TestCsv', '_TC.html'),
        'stub': (3, '', None, 'AMSTB_SrcFile.c')
    }

    def __init__(self, path, **kargs):
        self.path = Path(path)
        self.func = kargs.get('func')
        self.src_name = kargs['src_name']
        self.mode = self.get_mode()

    def collect_files(self):
        '''Get test result files path'''
        def relative_path(level, sub, ext, *args):
            fname = '{0}{1}'.format(self.func, ext) \
                if ext != None else args[0]

            return self.path.parents[level].joinpath(sub, fname)

        def modify_config(value):
            if self.mode == 'svn':
                value = [0, ''] + list(value[2:])

            return tuple(value)

        result = {key: relative_path(*modify_config(value))
                  for key, value in self.config.items()}
        result.update({'txt': self.path})

        if self.mode == 'svn':
            result.update({
                'xlsx': relative_path(2, '単体テスト仕様書', '.xlsx')
            })

        return result

    def get_mode(self):
        '''Get mode ams'''
        if self.path.parent.name == self.src_name:
            mode = 'ams'
        elif self.path.parent.name == self.func:
            mode = 'svn'
        else:
            raise Exception("Sorry boss ! I don't know what to do now")

        return mode
