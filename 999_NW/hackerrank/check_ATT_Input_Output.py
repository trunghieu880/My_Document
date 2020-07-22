# -*- coding: utf-8 -*-

import io
import logging
import sys
from pathlib import Path
import copy
import json

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

def get_xlsx_sheets(xlsx):
    '''Get sheets of xlsx'''
    logger.debug("Get sheets from file %s", xlsx)
    try:
        with open(xlsx, 'rb') as fp:
            xlsx = io.BytesIO(fp.read())

        wb = load_workbook(xlsx, read_only=True)
        data = wb.sheetnames
    except Exception as e:
        logger.exception(e)
        data = []
    finally:
        wb.close()
        return data

def get_xlsx_cells(xlsx, sheet, list_cell):
    '''Get cell value from excel file'''
    def val(cell):
        return str(cell.value) if cell.is_date else cell.value

    logger.debug("Get value of cell %s", list_cell)

    try:
        with open(xlsx, 'rb') as fp:
            xlsx = io.BytesIO(fp.read())

        wb = load_workbook(xlsx, read_only=True)
        sheet = sheet if isinstance(sheet, str) else wb.sheetnames[sheet-1]
        ws = wb[sheet]

        data = {key: val(ws[key]) for key in list_cell}
    except Exception as e:
        logger.exception(e)
        data = {}
    finally:
        wb.close()
        return data

def get_xlsx_raw(xlsx, sheet, begin=1, end=sys.maxsize, headers={}):
    '''Get raw data of table from excel.'''
    def val(cell):
        return str(cell.value) if cell.is_date else cell.value

    logger.debug("Get raw data from %s %s", Path(xlsx).name, sheet)

    try:
        with open(xlsx, 'rb') as fp:
            xlsx = io.BytesIO(fp.read())

        wb = load_workbook(xlsx, read_only=True, data_only=True)
        sheet = sheet if isinstance(sheet, str) else wb.sheetnames[sheet-1]

        data = [[val(cell) for cell in row] for row in wb[sheet].rows]
        data = data[begin-1:min(end, sys.maxsize)]

        first_row = data[0][:]
        data[0] = [headers.get(col, col) for col in data[0]]

        if headers != {}:
            data.append(first_row)

    except Exception as e:
        logger.exception(e)
        data = []
    finally:
        wb.close()
        return data

def check_flag(list_data, pat):
    result = False
    for x in list_data:
        if x == pat:
            result = True

    return result

def main():
    file = "C:\\Users\\hieu.nguyen-trung\\Desktop\\a.xlsx"
    lst_sheet = [x for x in get_xlsx_sheets(file)]
    data = get_xlsx_raw(file, lst_sheet[0], begin=22, end=22)
    # print(data)

    header = ['Inputs', 'Local variables as inputs', 'Imported Parameters', 'Local Parameters']
    list_flag = [check_flag(data[0], x) for x in header]

    '''
        Count element
    '''
    d = {'Inputs': 0, 'Local variables as inputs': 0, 'Imported Parameters': 0, 'Local Parameters': 0}

    for index, __header__ in enumerate(header):
        count = 0
        if list_flag[index]:
            flag_count = False
            temp_header = list(header)
            temp_header.remove(__header__)
            for x in data[0][data[0].index(__header__):]:
                if x in temp_header:
                    break
                if x == __header__ or (flag_count == True and x == None):
                    flag_count = True
                    count += 1
            d.update({__header__: count})

    # print(d)

    for key, value in d.items():
        if not(value > 0):
            d.delete(key)
    
    data = get_xlsx_raw(file, lst_sheet[0], begin=1, end=26)
    # print(data)
    temp_col_start = 1
    temp_col_end = 1 + d['Inputs'] - 1
    
    temp_col_start = 0
    temp_col_start = 0
    offset = d['Inputs']
    
    result = dict()

    template_obj = {
        "Tolerance": -1,
        "Type": None,
        "Max": -1,
        "Min": -1,
        "Data": list()
    }
    
    for __header__ in header:
        if __header__ == "Inputs":
            offset = 1
        else:
            offset = temp_col_end + 1
        
        temp_col_start = offset
        temp_col_end = temp_col_start + d[__header__] - 1
        
        # print(__header__)
        result[__header__] = list()
        for j in range(temp_col_start, temp_col_end + 1):
        
            name_input = data[22][j]
            data_js_variable = copy.deepcopy(template_obj)
            
            data_js_variable["Tolerance"] = data[17][j]
            data_js_variable["Type"] = data[18][j]
            data_js_variable["Max"] = data[19][j]
            data_js_variable["Min"] = data[20][j]
            
            for i in range(23, 26):
                data_js_variable["Data"].append(data[i][j])
            
            result[__header__].append({name_input: data_js_variable})

    # print(json.dumps(result, indent=2))
    
    for __header__ in header:
        print(result[__header__])
        
        for data_var in result[__header__]:
            for key, obj_level_2 in data_var.items():
                list_data_var = list(set(obj_level_2["Data"]))
                list_data_var.sort(key=None, reverse=False)
                # print(list_data_var)
                if (max(list_data_var)) > obj_level_2["Max"]:
                    print("OUT RANGE MAX")
                if (min(list_data_var)) < obj_level_2["Min"]:
                    print("OUT RANGE MIN")                    
        
    
main()
